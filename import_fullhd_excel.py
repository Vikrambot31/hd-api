#!/usr/bin/env python3
"""Import FullHD lookup table from Excel and optionally apply it.

Usage:
  python import_fullhd_excel.py --xlsx "path/to/file.xlsx" --sheet FullHD --out fullhd_lookup_new.json
  python import_fullhd_excel.py --xlsx "path/to/file.xlsx" --apply
"""

from __future__ import annotations

import argparse
import json
import shutil
from pathlib import Path


def fail(msg: str, code: int = 1) -> None:
    print(f"ERROR: {msg}")
    raise SystemExit(code)


def as_float(x):
    if x is None:
        return None
    try:
        return float(x)
    except Exception:
        return None


def as_int(x):
    if x is None:
        return None
    try:
        return int(x)
    except Exception:
        return None


def load_rows_from_excel(xlsx_path: Path, sheet_name: str):
    try:
        import openpyxl
    except Exception:
        fail("openpyxl not installed. Run: python -m pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        fail(f"sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")

    ws = wb[sheet_name]
    rows = []
    for r in ws.iter_rows(values_only=True):
        if not r:
            continue
        # Expected first 5 columns: lon_start, gate, line, color, tone
        lon = as_float(r[0] if len(r) > 0 else None)
        gate = as_int(r[1] if len(r) > 1 else None)
        line = as_int(r[2] if len(r) > 2 else None)
        color = as_int(r[3] if len(r) > 3 else None)
        tone = as_int(r[4] if len(r) > 4 else None)

        if lon is None or gate is None or line is None or color is None or tone is None:
            continue

        # Normalize longitude to [0,360)
        lon = lon % 360.0

        if not (1 <= gate <= 64 and 1 <= line <= 6 and 1 <= color <= 6 and 1 <= tone <= 6):
            continue

        rows.append([round(lon, 8), gate, line, color, tone])

    if not rows:
        fail("no valid rows found in Excel sheet")

    rows.sort(key=lambda x: x[0])
    return rows


def validate(rows):
    issues = []
    if len(rows) < 10000:
        issues.append(f"too few rows: {len(rows)}")

    lons = [r[0] for r in rows]
    if lons[0] < 0 or lons[-1] >= 360:
        issues.append(f"longitude range invalid: first={lons[0]}, last={lons[-1]}")

    gates = {r[1] for r in rows}
    if len(gates) != 64:
        issues.append(f"unique gates count is {len(gates)} (expected 64)")

    return issues


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Path to source Excel file")
    parser.add_argument("--sheet", default="FullHD", help="Sheet name (default: FullHD)")
    parser.add_argument("--out", default="fullhd_lookup_new.json", help="Output JSON path")
    parser.add_argument("--apply", action="store_true", help="Backup and replace fullhd_lookup.json")
    args = parser.parse_args()

    project_dir = Path(__file__).resolve().parent
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_absolute():
        xlsx_path = (project_dir / xlsx_path).resolve()

    if not xlsx_path.exists():
        fail(f"Excel file not found: {xlsx_path}")

    rows = load_rows_from_excel(xlsx_path, args.sheet)
    issues = validate(rows)

    out_path = Path(args.out)
    if not out_path.is_absolute():
        out_path = (project_dir / out_path).resolve()

    out_path.write_text(json.dumps(rows, separators=(",", ":"), ensure_ascii=False), encoding="utf-8")

    print(f"saved: {out_path}")
    print(f"rows: {len(rows)}")
    print(f"first: {rows[0][0]}  last: {rows[-1][0]}")

    if issues:
        print("WARNING: validation issues:")
        for i in issues:
            print(f"  - {i}")
    else:
        print("validation: OK")

    if args.apply:
        target = project_dir / "fullhd_lookup.json"
        backup = project_dir / "fullhd_lookup.backup.json"
        shutil.copy2(target, backup)
        shutil.copy2(out_path, target)
        print(f"applied: {target}")
        print(f"backup:  {backup}")


if __name__ == "__main__":
    main()
